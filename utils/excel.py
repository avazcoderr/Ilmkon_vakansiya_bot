"""
Professional Excel eksport — Ilmkon School arizalari.
openpyxl yordamida rang-barang, professional xlsx fayl yaratadi.
"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── Ranglar ──────────────────────────────────────────────────────────────────

BLUE_DARK   = "1F4E79"
BLUE_MID    = "2E75B6"
BLUE_LIGHT  = "D6E4F0"
BLUE_HEADER = "BDD7EE"
GREEN       = "375623"
GREEN_LIGHT = "E2EFDA"
RED_DARK    = "9C0006"
RED_LIGHT   = "FFC7CE"
AMBER_DARK  = "7D6608"
AMBER_LIGHT = "FFEB9C"
GRAY_LIGHT  = "F2F2F2"
WHITE       = "FFFFFF"

STATUS_COLORS = {
    "accepted":    (GREEN,     GREEN_LIGHT),
    "rejected":    (RED_DARK,  RED_LIGHT),
    "completed":   (AMBER_DARK, AMBER_LIGHT),
    "in_progress": ("595959",  GRAY_LIGHT),
    "cancelled":   ("808080",  "D9D9D9"),
}
STATUS_LABEL = {
    "accepted":    "✅ Qabul",
    "rejected":    "❌ Rad etildi",
    "completed":   "⏳ Kutilmoqda",
    "in_progress": "🔄 Jarayonda",
    "cancelled":   "🚫 Bekor",
}


def _border(color="BFBFBF", style="thin"):
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=10, color="000000", name="Calibri"):
    return Font(bold=bold, size=size, color=color, name=name)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Asosiy funksiya ──────────────────────────────────────────────────────────

def build_excel(rows: list[tuple]) -> bytes:
    wb = Workbook()

    # ── 1-varaq: Barcha arizalar ─────────────────────────────────────────────
    ws_all: Worksheet = wb.active
    ws_all.title = "Barcha arizalar"
    _build_all_sheet(ws_all, rows)

    # ── 2-varaq: Statistika ──────────────────────────────────────────────────
    ws_stat: Worksheet = wb.create_sheet("Statistika")
    _build_stats_sheet(ws_stat, rows)

    # ── 3-varaq: Qabul qilinganlar ────────────────────────────────────────────
    accepted = [(a, ans) for a, ans in rows if a.status == "accepted"]
    if accepted:
        ws_acc: Worksheet = wb.create_sheet("Qabul qilinganlar")
        _build_filtered_sheet(ws_acc, accepted, "✅ Qabul qilingan nomzodlar")

    # ── 4-varaq: Kutilmoqda ──────────────────────────────────────────────────
    pending = [(a, ans) for a, ans in rows if a.status == "completed"]
    if pending:
        ws_pend: Worksheet = wb.create_sheet("Kutilmoqda")
        _build_filtered_sheet(ws_pend, pending, "⏳ Ko'rib chiqilmagan arizalar")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 1-varaq: Barcha arizalar ─────────────────────────────────────────────────

def _build_all_sheet(ws: Worksheet, rows: list[tuple]) -> None:
    # Barcha field keylarini yig'ish
    all_keys: list[str] = []
    all_labels: dict[str, str] = {}
    for _, answers in rows:
        for ans in answers:
            if ans.field_key not in all_keys and ans.value_type == "text":
                all_keys.append(ans.field_key)
                all_labels[ans.field_key] = ans.field_label

    # Ustun ro'yxati
    fixed = ["№", "Ariza ID", "Lavozim", "Holat", "Sana", "Telegram ID"]
    all_cols = fixed + [all_labels.get(k, k) for k in all_keys]

    # ── Sarlavha qatori ──────────────────────────────────────────────────────
    for col_idx, header in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _font(bold=True, size=11, color=WHITE)
        cell.fill      = _fill(BLUE_DARK)
        cell.alignment = _align("center")
        cell.border    = _border(BLUE_MID)

    # ── Ma'lumot qatorlari ───────────────────────────────────────────────────
    for row_num, (app, answers) in enumerate(rows, 2):
        ans_dict = {a.field_key: a.value for a in answers if a.value_type == "text"}
        status_fg, status_bg = STATUS_COLORS.get(app.status, ("595959", GRAY_LIGHT))

        row_data = [
            row_num - 1,
            app.id,
            app.position_label,
            STATUS_LABEL.get(app.status, app.status),
            app.created_at.strftime("%d.%m.%Y %H:%M") if app.created_at else "",
            str(app.user_telegram_id),
        ] + [ans_dict.get(k, "") for k in all_keys]

        row_fill = _fill(BLUE_LIGHT) if row_num % 2 == 0 else _fill(WHITE)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border    = _border()
            cell.alignment = _align(wrap=True)

            # Status ustuni (4-ustun) — alohida rang
            if col_idx == 4:
                cell.font = _font(bold=True, color=status_fg)
                cell.fill = _fill(status_bg)
                cell.alignment = _align("center")
            else:
                cell.font = _font()
                cell.fill = row_fill

    # ── Ustun kengliklari ────────────────────────────────────────────────────
    col_widths = [5, 8, 24, 18, 18, 16] + [max(18, len(all_labels.get(k, k)) + 4) for k in all_keys]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(width, 40)

    # Birinchi qatorni muzlatish
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}1"
    ws.row_dimensions[1].height = 22

    # ── Jami satr ────────────────────────────────────────────────────────────
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="JAMI:")
    ws.cell(row=total_row, column=1).font = _font(bold=True, size=11)
    ws.cell(row=total_row, column=2, value=len(rows))
    ws.cell(row=total_row, column=2).font = _font(bold=True, size=11, color=BLUE_DARK)
    ws.cell(row=total_row, column=3, value=f"ta ariza  |  {datetime.now().strftime('%d.%m.%Y %H:%M')} holatida")
    ws.cell(row=total_row, column=3).font = _font(size=10, color="595959")
    for c in range(1, len(all_cols) + 1):
        ws.cell(row=total_row, column=c).fill = _fill(BLUE_LIGHT)


# ── 2-varaq: Statistika ──────────────────────────────────────────────────────

def _build_stats_sheet(ws: Worksheet, rows: list[tuple]) -> None:
    from data.questions import POSITIONS

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    def title(text, row):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font      = _font(bold=True, size=13, color=WHITE)
        cell.fill      = _fill(BLUE_DARK)
        cell.alignment = _align("center")
        ws.merge_cells(f"A{row}:C{row}")

    def header(label, val, row, bg=BLUE_LIGHT):
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=val)
        c1.font = _font(size=11); c1.fill = _fill(bg); c1.border = _border(); c1.alignment = _align()
        c2.font = _font(bold=True, size=11, color=BLUE_DARK); c2.fill = _fill(bg)
        c2.border = _border(); c2.alignment = _align("center")

    # Umumiy
    title("📊 UMUMIY STATISTIKA — ILMKON SCHOOL", 1)
    ws.row_dimensions[1].height = 28

    from datetime import date, timedelta
    today = date.today()
    today_cnt  = sum(1 for a, _ in rows if a.created_at and a.created_at.date() == today)
    week_start = today - timedelta(days=today.weekday())
    week_cnt   = sum(1 for a, _ in rows if a.created_at and a.created_at.date() >= week_start)
    month_cnt  = sum(1 for a, _ in rows if a.created_at and a.created_at.month == today.month and a.created_at.year == today.year)

    header("📦  Jami arizalar",       len(rows),    2)
    header("📅  Bugun",               today_cnt,    3)
    header("📆  Bu hafta",            week_cnt,     4)
    header("🗓  Bu oy",               month_cnt,    5)

    ws.cell(row=6, column=1).value = ""

    # Holat bo'yicha
    title("📈 HOLAT BO'YICHA", 7)
    ws.row_dimensions[7].height = 22

    status_rows = [
        ("⏳  Kutilmoqda (completed)", "completed", AMBER_LIGHT),
        ("✅  Qabul qilindi",          "accepted",  GREEN_LIGHT),
        ("❌  Rad etildi",             "rejected",  RED_LIGHT),
        ("🔄  Jarayonda",              "in_progress", GRAY_LIGHT),
        ("🚫  Bekor qilindi",          "cancelled", "EBEBEB"),
    ]
    for i, (label, status, bg) in enumerate(status_rows, 8):
        cnt = sum(1 for a, _ in rows if a.status == status)
        pct = f"{cnt/len(rows)*100:.1f}%" if rows else "0%"
        c1 = ws.cell(row=i, column=1, value=label)
        c2 = ws.cell(row=i, column=2, value=cnt)
        c3 = ws.cell(row=i, column=3, value=pct)
        for c in (c1, c2, c3):
            c.fill = _fill(bg); c.border = _border(); c.font = _font(size=11)
        c2.font = _font(bold=True, size=11); c2.alignment = _align("center")
        c3.alignment = _align("center")

    ws.cell(row=13, column=1).value = ""

    # Lavozim bo'yicha
    title("💼 LAVOZIM BO'YICHA", 14)
    ws.row_dimensions[14].height = 22

    fills_cycle = [BLUE_LIGHT, "E8F4FD", BLUE_LIGHT, "E8F4FD", BLUE_LIGHT]
    for i, (key, pos_label) in enumerate(POSITIONS.items(), 15):
        cnt  = sum(1 for a, _ in rows if a.position_key == key)
        pct  = f"{cnt/len(rows)*100:.1f}%" if rows else "0%"
        acc  = sum(1 for a, _ in rows if a.position_key == key and a.status == "accepted")
        bg   = fills_cycle[i - 15]
        c1 = ws.cell(row=i, column=1, value=pos_label)
        c2 = ws.cell(row=i, column=2, value=cnt)
        c3 = ws.cell(row=i, column=3, value=f"{pct}  (✅ {acc} qabul)")
        for c in (c1, c2, c3):
            c.fill = _fill(bg); c.border = _border(); c.font = _font(size=11)
        c2.font = _font(bold=True, size=11, color=BLUE_DARK); c2.alignment = _align("center")

    ws.cell(row=20, column=1).value = ""

    # Oxirgi 7 kun
    title("📅 OXIRGI 7 KUN", 21)
    ws.row_dimensions[21].height = 22

    from datetime import timedelta
    for i in range(7):
        d = today - timedelta(days=6 - i)
        cnt = sum(1 for a, _ in rows if a.created_at and a.created_at.date() == d)
        c1 = ws.cell(row=22 + i, column=1, value=d.strftime("%d.%m.%Y (%A)"))
        c2 = ws.cell(row=22 + i, column=2, value=cnt)
        bar = "█" * cnt + "░" * max(0, 5 - cnt)
        c3 = ws.cell(row=22 + i, column=3, value=bar)
        bg = BLUE_LIGHT if i % 2 == 0 else WHITE
        for c in (c1, c2, c3):
            c.fill = _fill(bg); c.border = _border(); c.font = _font(size=10)
        c2.font = _font(bold=True, size=11, color=BLUE_DARK); c2.alignment = _align("center")
        c3.font = Font(name="Courier New", size=10, color=BLUE_MID)

    # Hisobot sanasi
    ws.cell(row=30, column=1, value=f"Hisobot sanasi: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    ws.cell(row=30, column=1).font = _font(size=9, color="808080")


# ── Filtrlangan varaq ────────────────────────────────────────────────────────

def _build_filtered_sheet(ws: Worksheet, rows: list[tuple], title: str) -> None:
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 28

    # Sarlavha
    hdr = ws.cell(row=1, column=1, value=title)
    hdr.font = _font(bold=True, size=12, color=WHITE)
    hdr.fill = _fill(BLUE_DARK)
    hdr.alignment = _align("center")
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 26

    # Ustun sarlavhalari
    cols = ["№", "ID", "Lavozim", "Holat", "Ism Familiya", "Telefon", "Sana"]
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=2, column=i, value=c)
        cell.font = _font(bold=True, size=10, color=WHITE)
        cell.fill = _fill(BLUE_MID)
        cell.border = _border()
        cell.alignment = _align("center")

    for row_num, (app, answers) in enumerate(rows, 3):
        ans_dict = {a.field_key: a.value for a in answers if a.value_type == "text"}
        bg = GREEN_LIGHT if app.status == "accepted" else AMBER_LIGHT
        vals = [
            row_num - 2,
            app.id,
            app.position_label,
            STATUS_LABEL.get(app.status, app.status),
            ans_dict.get("fullname", "—"),
            ans_dict.get("phone", "—"),
            app.created_at.strftime("%d.%m.%Y %H:%M") if app.created_at else "—",
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=ci, value=v)
            cell.font   = _font(size=10)
            cell.fill   = _fill(bg if row_num % 2 == 0 else WHITE)
            cell.border = _border()
            cell.alignment = _align(wrap=True)

    ws.freeze_panes = "A3"
